"""Excel 成绩单导入（DESIGN §5 P0 入口）。

格式约定（第一行为表头）：
    姓名 | Q1 | Q2 | ... | QN | 总分(可选)
- 第 1 列按 name_or_alias 或 external_code 匹配学生；
- 题列按表头中的题号匹配（Q1 / 第1题 / 1 均可）；
- 未填分数按 0 分处理并记入 warnings（由教师在审核阶段核对）；
- 导入产物状态 = 待审核（状态机隔离，不变量①）。
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import (
    ExamResponse,
    ExamTemplate,
    ResponseAnswer,
    Student,
    TemplateQuestion,
)


@dataclass
class ImportResult:
    imported: int = 0
    unmatched_students: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def import_excel(
    session: Session, template_id: int, file_path: str | Path
) -> ImportResult:
    template = session.get(ExamTemplate, template_id)
    if template is None:
        raise ValueError(f"exam_template {template_id} 不存在")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValueError("Excel 为空")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    result = ImportResult()

    name_col = _find_name_column(header)
    total_col = _find_total_column(header)
    q_cols = _map_question_columns(session, template, header, result)

    students = {
        s.name_or_alias: s
        for s in session.scalars(
            select(Student).where(Student.class_id == template.class_id)
        )
    }
    students.update(
        {
            s.external_code: s
            for s in session.scalars(
                select(Student).where(
                    Student.class_id == template.class_id,
                    Student.external_code != "",
                )
            )
        }
    )

    for r_idx, row in enumerate(rows[1:], start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        raw_name = str(row[name_col]).strip() if row[name_col] is not None else ""
        student = students.get(raw_name)
        if student is None:
            result.unmatched_students.append(f"第{r_idx}行:{raw_name}")
            continue

        existing = session.scalar(
            select(ExamResponse.id).where(
                ExamResponse.exam_template_id == template.id,
                ExamResponse.student_id == student.id,
            )
        )
        if existing is not None:
            result.warnings.append(f"第{r_idx}行:{raw_name} 已有作答记录，跳过")
            continue

        response = ExamResponse(
            exam_template_id=template.id,
            student_id=student.id,
            source="excel",
            status="待审核",
        )
        session.add(response)
        session.flush()

        computed_total = 0.0
        for tq, col in q_cols.items():
            raw = row[col] if col < len(row) else None
            if raw is None or str(raw).strip() == "":
                result.warnings.append(
                    f"第{r_idx}行:{raw_name} 第{tq.idx}题未填，按 0 分处理"
                )
                score = 0.0
            else:
                score = _parse_score(raw, tq.full_score, raw_name, tq.idx, result)
            computed_total += score
            session.add(
                ResponseAnswer(
                    exam_response_id=response.id,
                    template_question_id=tq.id,
                    score=score,
                )
            )

        declared_total = None
        if total_col is not None and total_col < len(row) and row[total_col] is not None:
            try:
                declared_total = float(row[total_col])
            except (TypeError, ValueError):
                result.warnings.append(f"第{r_idx}行:{raw_name} 总分列无法解析")
        if declared_total is not None and abs(declared_total - computed_total) > 0.01:
            result.warnings.append(
                f"第{r_idx}行:{raw_name} 总分 {declared_total} 与逐题求和 "
                f"{computed_total} 不一致，以求和为准"
            )
        response.total_score = round(computed_total, 2)
        result.imported += 1

    session.flush()
    return result


def _find_name_column(header: list[str]) -> int:
    for i, h in enumerate(header):
        if h in ("姓名", "学生", "学生姓名", "名字", "学号", "name", "Name"):
            return i
    return 0  # 默认第一列


def _find_total_column(header: list[str]) -> int | None:
    for i, h in enumerate(header):
        if h in ("总分", "总成绩", "total", "Total"):
            return i
    return None


def _map_question_columns(
    session: Session, template: ExamTemplate, header: list[str], result: ImportResult
) -> dict[TemplateQuestion, int]:
    """表头题号 → template_question 映射。支持 Q1 / 第1题 / 1。"""
    questions = {
        q.idx: q
        for q in session.scalars(
            select(TemplateQuestion).where(
                TemplateQuestion.exam_template_id == template.id
            )
        )
    }
    mapping: dict[TemplateQuestion, int] = {}
    for i, h in enumerate(header):
        m = re.search(r"(\d+)", h)
        if not m:
            continue
        idx = int(m.group(1))
        if idx in questions and questions[idx] not in mapping:
            mapping[questions[idx]] = i
    missing = [idx for idx in questions if questions[idx] not in mapping]
    if missing:
        result.warnings.append(f"模板题目 {sorted(missing)} 在 Excel 表头中无对应列")
    return mapping


def _parse_score(
    raw, full_score: float, student_name: str, q_idx: int, result: ImportResult
) -> float:
    try:
        score = float(raw)
    except (TypeError, ValueError):
        result.warnings.append(
            f"{student_name} 第{q_idx}题分数无法解析：{raw!r}，按 0 分处理"
        )
        return 0.0
    if score < 0 or score > full_score:
        result.warnings.append(
            f"{student_name} 第{q_idx}题分数越界：{score}（满分 {full_score}），已裁剪"
        )
        score = max(0.0, min(full_score, score))
    return score
